import os
import smtplib
from email.message import EmailMessage
import pandas as pd
import openpyxl
import io
import uuid
import unicodedata
from datetime import date
import pythoncom
import win32com.client
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from io import BytesIO
from dotenv import load_dotenv
import psycopg
from psycopg.rows import dict_row
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

load_dotenv()

app = Flask(__name__)
CORS(
    app,
    resources={r"/*": {"origins": ["http://127.0.0.1:5500", "http://localhost:5500", "http://127.0.0.1:5000", "http://localhost:5000"]}},
    supports_credentials=True,
    allow_headers=["Content-Type", "Authorization"],
    methods=["GET", "POST", "OPTIONS"]
)

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_PASSWORD = os.getenv("GMAIL_PASSWORD")

print(
    "GMAIL:",
    GMAIL_USER
)

history_store = []

try:
    conn = psycopg.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        row_factory=dict_row
    )
except Exception as exc:
    conn = None
    print(f"No se pudo conectar a PostgreSQL: {exc}")


def ensure_schema():
    if conn is None:
        return

    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historico_budget (
                id SERIAL PRIMARY KEY,
                periodo TEXT,
                sgi TEXT,
                nombre_completo TEXT,
                bu TEXT,
                tipo_empleado TEXT,
                cco TEXT,
                puesto TEXT,
                area TEXT,
                fecha_ingreso TEXT,
                antiguedad TEXT,
                fecha_nacimiento TEXT,
                edad TEXT,
                genero TEXT,
                salario_diario TEXT,
                salario_mensual TEXT,
                inc_salarial TEXT,
                promocion TEXT,
                nivelacion TEXT,
                suma_porcentual TEXT,
                nuevo_salario TEXT,
                comentarios TEXT,
                manager TEXT,
                fecha_registro TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        conn.commit()


ensure_schema()


def normalize_text(value):
    return '' if value is None else str(value).strip()


def to_number(value):
    if value in (None, ''):
        return None

    try:
        return float(str(value).replace('%', '').replace('$', '').replace(',', '').strip())
    except (TypeError, ValueError):
        return None


def to_number(valor):

    if valor is None:
        return 0

    if valor == "":
        return 0

    try:
        return float(valor)
    except:
        return 0

def to_date(valor):

    if valor is None:
        return None

    if valor == "":
        return None

    return valor        

def build_history_row(record, manager, periodo):
    inc = to_number(record.get('incSalarial', record.get('inc_salarial')))
    promocion = to_number(record.get('promocion'))
    nivelacion = to_number(record.get('nivelacion'))
    salario_mensual = to_number(record.get('sm', record.get('salario_mensual')))

    suma_porcentual = None
    nuevo_salario = None

    if inc is not None or promocion is not None or nivelacion is not None:
        values = [value for value in [inc, promocion, nivelacion] if value is not None]
        suma_porcentual = sum(values) if values else 0

    if salario_mensual is not None and suma_porcentual is not None:
        nuevo_salario = salario_mensual + (salario_mensual * suma_porcentual / 100)

    return {
        'periodo': periodo,
        'sgi': normalize_text(record.get('sgi')) or f"MANUAL-{uuid.uuid4().hex[:8]}",
        'nombre_completo': normalize_text(record.get('nombre')) or 'Registro Manual',
        'bu': normalize_text(record.get('bu')),
        'tipo_empleado': normalize_text(record.get('tipoEmpleado', record.get('tipo_empleado'))),
        'cco': normalize_text(record.get('cco')),
        'puesto': normalize_text(record.get('puesto')),
        'area': normalize_text(record.get('area')),
        'fecha_ingreso': to_date(record.get('fechaIngreso', record.get('fecha_ingreso'))),
        'antiguedad': normalize_text(record.get('antiguedad')),
        'fecha_nacimiento': to_date(record.get('fechaNacimiento', record.get('fecha_nacimiento'))),
        'edad': int(to_number(record.get('edad'))),
        'genero': normalize_text(record.get('genero')),
        'salario_diario': to_number(record.get('sd',record.get('salario_diario'))),
        'salario_mensual': to_number(record.get('sm', record.get('salario_mensual'))),
        'inc_salarial': inc,
        'promocion': promocion,
        'nivelacion': nivelacion,
        'suma_porcentual': suma_porcentual,
        'nuevo_salario': nuevo_salario,
        'comentarios': normalize_text(record.get('comentarios')),
        'manager': normalize_text(manager or record.get('manager') or 'General')
    }


def fetch_empleados():
    if conn is None:
        return []

    with conn.cursor() as cur:
        cur.execute("""
            SELECT *
            FROM empleados_sg
        """)
        return cur.fetchall()


def fetch_equipo(manager):
    if conn is None:
        return []

    with conn.cursor() as cur:
        cur.execute("""
            SELECT *
            FROM empleados_sg
            WHERE manager = %s
        """, (manager,))
        return cur.fetchall()

def generar_excel_manager(manager):

    registros = fetch_equipo(manager)

    if not registros:
        return None

    df = pd.DataFrame(registros)

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Budget",
            index=False
        )

    output.seek(0)

    return output

def generar_excel_manager_archivo(manager):

    registros = fetch_equipo(manager)

    print(registros[0])

    if not registros:
        return None

    normalizados = []

    for empleado in registros:

        normalizados.append({

            "SGI":
                empleado.get("sgi")
                or empleado.get("id"),

            "Nombre":
                empleado.get("nombre_completo")
                or empleado.get("nombre"),

            "BU":
                empleado.get("compania")
                or empleado.get("bu"),

            "Tipo Empleado":
                empleado.get("tipo_empleado"),

            "CCO":
                empleado.get("ceco"),

            "Puesto":
                empleado.get("puesto"),

            "Área":
                empleado.get("filiere"),

            "Fecha Ingreso":
                empleado.get("fecha_ingreso"),

            "Antigüedad":
                round(
                    (
                        date.today()
                        - empleado["fecha_ingreso"]
                    ).days / 365.25,
                    1
                )
                if empleado.get("fecha_ingreso")
                else "",

            "Fecha Nacimiento":
                empleado.get("fecha_nacimiento"),

            "Edad":
                int(
                    (
                        date.today()
                        - empleado["fecha_nacimiento"]
                    ).days / 365.25
                )
                if empleado.get("fecha_nacimiento")
                else "",

            "Genero":
                empleado.get("sexo"),

            "Salario Diario":
                float(
                    empleado["salario_diario"]
                )
                if empleado.get("salario_diario")
                else "",

            "Salario Mensual":
                float(
                    empleado["salario_mensual"]
                )
                if empleado.get("salario_mensual")
                else "",

            "Incremento Salarial":
                empleado.get("inc_salarial"),

            "Promoción":
                empleado.get("promocion"),

            "Nivelación":
                empleado.get("nivelacion"),

            "Comentarios":
                empleado.get("comentarios"),
        })

    df = pd.DataFrame(normalizados)

    df = df[[
        "SGI",
        "Nombre",
        "BU",
        "Tipo Empleado",
        "CCO",
        "Puesto",
        "Área",
        "Fecha Ingreso",
        "Antigüedad",
        "Fecha Nacimiento",
        "Edad",
        "Genero",
        "Salario Diario",
        "Salario Mensual",
        "Incremento Salarial",
        "Promoción",
        "Nivelación",
        "Comentarios",
    ]]

    carpeta_temp = "temp_excels"

    os.makedirs(
        carpeta_temp,
        exist_ok=True
    )

    file_path = os.path.join(
        carpeta_temp,
        f"Budget_{manager}.xlsx"
            .replace(" ", "_")
    )

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Budget",
            index=False
        )

        ws = writer.sheets["Budget"]

        # ==========================================
        # ESTILOS
        # ==========================================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="005EB8"
        )

        editable_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

        formula_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        white_font = Font(
            color="FFFFFF",
            bold=True
        )

        # ==========================================
        # ENCABEZADOS
        # ==========================================

        ws["R1"] = "Suma %"
        ws["S1"] = "Nuevo Salario"

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = white_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Encabezados calculados

        ws["R1"].fill = formula_fill
        ws["S1"].fill = formula_fill

        ws["R1"].font = Font(
            bold=True
        )

        ws["S1"].font = Font(
            bold=True
        )

        ws["R1"].alignment = Alignment(
            horizontal="center"
        )

        ws["S1"].alignment = Alignment(
            horizontal="center"
)

        # ==========================================
        # FILAS
        # ==========================================

        for row in range(
            2,
            len(df) + 2
        ):

            # Columnas capturables

            ws[f"O{row}"].fill = editable_fill

            ws[f"P{row}"].fill = editable_fill

            ws[f"Q{row}"].fill = editable_fill

            # Columnas calculadas

            ws[f"R{row}"].fill = formula_fill

            ws[f"S{row}"].fill = formula_fill

            # Fórmula suma %

            ws[f"R{row}"] = (
                f"=O{row}+P{row}+Q{row}"
            )

            # Fórmula nuevo salario

            ws[f"S{row}"] = (
                f"=N{row}*(1+(R{row}/100))"
            )

            # Formato moneda

            ws[f"M{row}"].number_format = '$#,##0.00'

            ws[f"N{row}"].number_format = '$#,##0.00'

            ws[f"S{row}"].number_format = '$#,##0.00'

        # ==========================================
        # CONGELAR ENCABEZADO
        # ==========================================

        ws.freeze_panes = "A2"

        # ==========================================
        # FILTROS
        # ==========================================

        ws.auto_filter.ref = ws.dimensions

        # ==========================================
        # AUTO AJUSTE COLUMNAS
        # ==========================================

        for column in ws.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = max_length + 3

    return os.path.abspath(
        file_path
    )

def fetch_managers():
    if conn is None:
        return []

    with conn.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT manager
            FROM empleados_sg
            WHERE manager IS NOT NULL
              AND manager <> COALESCE(sgi, '')
            ORDER BY manager
        """)
        return cur.fetchall()


@app.route("/")
def home():
    return jsonify({
        "mensaje": "Backend funcionando"
    })


@app.route("/empleados")
def empleados():
    return jsonify(fetch_empleados())


@app.route("/equipo/<manager>")
def equipo(manager):
    return jsonify(fetch_equipo(manager))


@app.route("/managers")
def managers():
    return jsonify(fetch_managers())


@app.route("/test_outlook")
def test_outlook():

    try:

        pythoncom.CoInitialize()

        outlook = win32com.client.Dispatch(
            "Outlook.Application"
        )

        mail = outlook.CreateItem(0)

        mail.To = "TU_CORREO_PRUEBA@gmail.com"

        mail.Subject = "Prueba"

        mail.Body = "Prueba"

        mail.Send()

        return jsonify({
            "ok": True,
            "message":
                "Correo enviado desde Outlook"
        })

    except Exception as e:

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    finally:

        pythoncom.CoUninitialize()


@app.route("/enviar_correos", methods=["POST"])
def enviar_correos():

    import pythoncom
    import win32com.client

    data = request.get_json()

    managers = data.get(
        "managers",
        []
    )

    resultado = []

    with conn.cursor() as cur:

        for manager in managers:

            cur.execute("""
                SELECT DISTINCT
                    nombre_completo,
                    correo
                FROM empleados_sg
                WHERE nombre_completo = %s
                LIMIT 1
            """, (manager,))

            row = cur.fetchone()

            if row:

                resultado.append({
                    "manager":
                        row["nombre_completo"],
                    "correo":
                        row["correo"]
                })

    print("\nCORREOS ENCONTRADOS:")
    print(resultado)

    pythoncom.CoInitialize()

    try:

        outlook = win32com.client.Dispatch(
            "Outlook.Application"
        )

        for item in resultado:

            mail = outlook.CreateItem(0)

            image_path = os.path.abspath(
                os.path.join(
                    "..",
                    "imagenes",
                    "correo.png"
                )
            )

            mail.To = item["correo"]

            mail.CC = (
                "norma.mendoza@saint-gobain.com;"
                "leyda.segundo@saint-gobain.com"
            )

            mail.Subject = (
                "Budget Planning 2027 - Revisión de Incrementos"
            )

            img = mail.Attachments.Add(
                image_path
            )

            img.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
                "correo_img"
            )

            mail.HTMLBody = f"""
            <html>

            <body style="font-family:Calibri, Arial, sans-serif; font-size:11pt; color:#000000;">

            <p>
            <strong>Estimado(a) {item["manager"]},</strong>
            </p>

            <p>
            Esperamos que se encuentre muy bien.
            </p>

            <p>
            Por medio del presente, compartimos la información correspondiente al presupuesto salarial del personal a su cargo, así como los importes considerados para el proceso de <strong>Budget 2027</strong>.
            </p>

            <p>
            Asimismo, <strong>les solicitamos considerar</strong> que el incremento presupuestal autorizado para dicho ejercicio corresponde a <strong>X.X%</strong>.
            </p>

            <p>
            <strong>Agradecemos nos puedan compartir</strong> su propuesta en los diferentes conceptos en caso de aplicar:
            </p>

            <ol>
                <li><strong>Incremento salarial</strong></li>
                <li><strong>Promoción</strong></li>
                <li><strong>Nivelación</strong></li>
            </ol>

            <p>
            Así como cualquier comentario u observación que consideren pertinente, a fin de incorporarlos al análisis y continuar con el proceso de autorización correspondiente.
            </p>

            <p>
            Quedamos atentos a sus comentarios y agradecemos de antemano su apoyo y valiosa colaboración.
            </p>

            <p>
            Muchas gracias.
            </p>

            <p>
            Saludos cordiales,
            </p>

            <p>
            <strong>Lázaro García</strong><br>
            Total Rewards
            </p>

            <br>

            <img src="cid:correo_img">

            </body>

            </html>
            """

            excel_path = generar_excel_manager_archivo(
                item["manager"]
            )

            print("\nMANAGER:")
            print(item["manager"])

            print("\nRUTA:")
            print(excel_path)

            print("\nEXISTE:")
            print(os.path.exists(excel_path))

            if excel_path:

                print("\nTAMAÑO:")
                print(os.path.getsize(excel_path))
                
                try:

                    mail.Attachments.Add(
                        excel_path
                    )

                    print("ADJUNTO OK")

                except Exception as e:

                    print("ERROR ADJUNTANDO:")
                    print(e)

                    raise

            mail.Display()

        return jsonify({
            "ok": True,
            "managers": resultado
        })

    finally:

        pythoncom.CoUninitialize()
#@app.route("/test_mail")
#def test_mail():
#
 #   print("ANTES SMTP")

 #   smtp = smtplib.SMTP(
  #      "smtp.gmail.com",
   #     587,
    #    timeout=20
    #)

    #print("DESPUES SMTP")

    #smtp.quit()

    #return jsonify({
     #   "ok": True
    #})
    
@app.route("/guardar_registro", methods=["GET", "POST", "OPTIONS"])
def guardar_registro():

    print("ENTRO A GUARDAR")

    if request.method == "OPTIONS":
        response = jsonify({"ok": True})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        return response

    payload = request.get_json(silent=True) or {}

    print("PAYLOAD:")
    print(payload)

    manager = payload.get("manager") or "General"
    periodo = payload.get("periodo") or "2027"
    registros = payload.get("registros") or []

    if not isinstance(registros, list):
        return jsonify({
            "ok": False,
            "error": "Formato inválido"
        }), 400

    try:

        if conn is not None:

            lote_id = str(uuid.uuid4())

            with conn.cursor() as cur:

                for record in registros:

                    row = build_history_row(
                        record,
                        manager,
                        periodo
                    )

                    print("\n--------------------")
                    print("ROW A GUARDAR:")
                    print(row)
                    print("--------------------\n")

                    cur.execute("""
                        INSERT INTO historico_budget (
                            lote_id,
                            periodo,
                            sgi,
                            nombre_completo,
                            bu,
                            tipo_empleado,
                            cco,
                            puesto,
                            area,
                            fecha_ingreso,
                            antiguedad,
                            fecha_nacimiento,
                            edad,
                            genero,
                            salario_diario,
                            salario_mensual,
                            inc_salarial,
                            promocion,
                            nivelacion,
                            suma_porcentual,
                            nuevo_salario,
                            comentarios,
                            manager,
                            aprobado,
                            fecha_registro
                        )
                        VALUES (
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, NOW()
                        )
                    """, (
                        lote_id,
                        row['periodo'],
                        row['sgi'],
                        row['nombre_completo'],
                        row['bu'],
                        row['tipo_empleado'],
                        row['cco'],
                        row['puesto'],
                        row['area'],
                        row['fecha_ingreso'],
                        row['antiguedad'],
                        row['fecha_nacimiento'],
                        row['edad'],
                        row['genero'],
                        row['salario_diario'],
                        row['salario_mensual'],
                        row['inc_salarial'],
                        row['promocion'],
                        row['nivelacion'],
                        row['suma_porcentual'],
                        row['nuevo_salario'],
                        row['comentarios'],
                        row['manager'],
                        False
                    ))

            conn.commit()

        else:

            history_store.append({
                "manager": manager,
                "timestamp": datetime.utcnow().isoformat(),
                "registros": registros
            })

        return jsonify({
            "ok": True,
            "message": "Cambios guardados correctamente",
            "manager": manager,
            "registros": len(registros)
        })

    except Exception as e:

        conn.rollback()

        print("\n==============================")
        print("ERROR REAL EN GUARDADO")
        print("==============================")
        print(str(e))
        print("==============================\n")

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/historial_detalle/<lote_id>")
def historial_detalle(lote_id):

    with conn.cursor() as cur:

        cur.execute("""
            SELECT *
            FROM historico_budget
            WHERE lote_id = %s
            ORDER BY nombre_completo
        """,(lote_id,))

        rows = cur.fetchall()

    return jsonify(rows)

@app.route("/historial/<manager>")
def historial(manager):
    if conn is not None:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT lote_id,manager, fecha_registro, aprobado, agregado_sabana, id, sgi, nombre_completo, bu, tipo_empleado, cco, puesto, area,
                       fecha_ingreso, antiguedad, fecha_nacimiento, edad, genero, salario_diario,
                       salario_mensual, inc_salarial, promocion, nivelacion, suma_porcentual,
                       nuevo_salario, comentarios
                FROM historico_budget
                WHERE manager = %s
                ORDER BY fecha_registro DESC, id DESC
            """, (manager,))
            rows = cur.fetchall()

        if rows:
            grouped = []
            for row in rows:
                timestamp_value = row['fecha_registro'].isoformat() if row['fecha_registro'] else datetime.utcnow().isoformat()
                if not grouped or grouped[-1]["lote_id"] != row["lote_id"]:
                    grouped.append({
                        "lote_id": row["lote_id"],
                        "manager": row["manager"],
                        "timestamp": timestamp_value,
                        "aprobado": row["aprobado"],
                        "agregado_sabana": row["agregado_sabana"],
                        "registros": []
                    })

                grouped[-1]['registros'].append({
                    'sgi': row['sgi'],
                    'nombre': row['nombre_completo'],
                    'bu': row['bu'],
                    'tipo_empleado': row['tipo_empleado'],
                    'cco': row['cco'],
                    'puesto': row['puesto'],
                    'area': row['area'],
                    'fecha_ingreso': row['fecha_ingreso'],
                    'antiguedad': row['antiguedad'],
                    'fecha_nacimiento': row['fecha_nacimiento'],
                    'edad': row['edad'],
                    'genero': row['genero'],
                    'salario_diario': row['salario_diario'],
                    'salario_mensual': row['salario_mensual'],
                    'inc_salarial': row['inc_salarial'],
                    'promocion': row['promocion'],
                    'nivelacion': row['nivelacion'],
                    'suma_porcentual': row['suma_porcentual'],
                    'nuevo_salario': row['nuevo_salario'],
                    'comentarios': row['comentarios']
                })

            return jsonify(grouped[:10])

    items = [item for item in history_store if item.get("manager") == manager]
    return jsonify(items)

@app.route("/excel/<lote_id>")
def descargar_excel(lote_id):

    with conn.cursor() as cur:

        cur.execute("""
            SELECT
                sgi,
                nombre_completo,
                bu,
                tipo_empleado,
                cco,
                puesto,
                area,
                fecha_ingreso,
                antiguedad,
                fecha_nacimiento,
                edad,
                genero,
                salario_diario,
                salario_mensual,
                inc_salarial,
                promocion,
                nivelacion,
                suma_porcentual,
                nuevo_salario,
                comentarios,
                manager
            FROM historico_budget
            WHERE lote_id = %s
            ORDER BY nombre_completo
        """, (lote_id,))

        rows = cur.fetchall()

    if not rows:

        return jsonify({
            "ok": False,
            "message": "No se encontraron registros"
        }), 404

    df = pd.DataFrame(rows)

    manager_name = (
        rows[0].get("manager")
        if rows
        else "Manager"
    )

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Budget",
            index=False
        )

    output.seek(0)

    file_name = (
        f"Budget_{manager_name}.xlsx"
        .replace(" ", "_")
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/aprobar/<lote_id>", methods=["POST"])
def aprobar_lote(lote_id):

    with conn.cursor() as cur:

        cur.execute("""
            UPDATE historico_budget
            SET aprobado = TRUE
            WHERE lote_id = %s
        """, (lote_id,))

    conn.commit()

    return jsonify({
        "ok": True,
        "message": "Histórico aprobado"
    })    

@app.route("/excel_actual", methods=["POST"])
def descargar_excel_actual():

    data = request.get_json()

    manager = data.get(
        "manager",
        "General"
    )

    registros = data.get(
        "registros",
        []
    )

    if not registros:

        return jsonify({
            "ok": False,
            "message": "No hay registros para exportar."
        }), 400

    df = pd.DataFrame(registros)

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Budget",
            index=False
        )

    output.seek(0)

    file_name = (
        f"Budget_{manager}.xlsx"
        .replace(" ", "_")
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/completados")
def completados():

    with conn.cursor() as cur:

        cur.execute("""
            SELECT DISTINCT
                lote_id,
                manager,
                fecha_registro,
                aprobado,
                agregado_sabana
            FROM historico_budget
            WHERE aprobado = TRUE
            AND agregado_sabana = FALSE
            ORDER BY fecha_registro DESC
        """)

        rows = cur.fetchall()

    return jsonify(rows)

@app.route("/preview_envio", methods=["POST"])
def preview_envio():

    data = request.get_json()

    managers = data.get(
        "managers",
        []
    )

    resultado = []

    with conn.cursor() as cur:

        for manager in managers:

            cur.execute("""
                SELECT DISTINCT
                    nombre_completo,
                    correo
                FROM empleados_sg
                WHERE nombre_completo = %s
                LIMIT 1
            """, (manager,))

            row = cur.fetchone()

            if row:

                resultado.append({
                    "manager":
                        row["nombre_completo"],
                    "correo":
                        row["correo"]
                })

    print("\nCORREOS ENCONTRADOS:")
    print(resultado)

    return jsonify({
        "ok": True,
        "managers": resultado
    })

@app.route("/excel_manager/<manager>")
def excel_manager(manager):

    output = generar_excel_manager(
        manager
    )

    if output is None:

        return jsonify({
            "ok": False
        }), 404

    file_name = (
        f"Budget_{manager}.xlsx"
        .replace(" ", "_")
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route(
    "/consolidar_excels",
    methods=["POST"]
)
def consolidar_excels():

    archivos = request.files.getlist(
        "files"
    )

    dfs = []

    for archivo in archivos:

        df = pd.read_excel(
            archivo,
            engine="openpyxl"
        )

        dfs.append(df)

    consolidado = pd.concat(
        dfs,
        ignore_index=True
    )

    archivos_leidos = len(
        archivos
    )

    registros_encontrados = len(
        consolidado
    )

    sgi_unicos = consolidado[
        "SGI"
    ].nunique()

    sgi_duplicados = (
        registros_encontrados
        - sgi_unicos
    )

    os.makedirs(
        "temp_excels",
        exist_ok=True
    )

    ruta_consolidado = os.path.join(
        "temp_excels",
        "Budget_Consolidado.xlsx"
    )

    with pd.ExcelWriter(
        ruta_consolidado,
        engine="openpyxl"
    ) as writer:

        consolidado.to_excel(
            writer,
            sheet_name="Consolidado",
            index=False
        )

    return jsonify({

        "ok": True,

        "archivosLeidos":
            archivos_leidos,

        "registrosEncontrados":
            registros_encontrados,

        "sgiUnicos":
            sgi_unicos,

        "sgiDuplicados":
            sgi_duplicados

    })

@app.route(
    "/descargar_consolidado"
)
def descargar_consolidado():

    ruta = os.path.join(
        "temp_excels",
        "Budget_Consolidado.xlsx"
    )

    if not os.path.exists(ruta):

        return jsonify({
            "ok": False,
            "message":
                "No existe consolidado disponible."
        }), 404

    return send_file(
        ruta,
        as_attachment=True,
        download_name=
            "Budget_Consolidado.xlsx"
    )

@app.route(
    "/actualizar_sabana",
    methods=["POST"]
)
def actualizar_sabana():
    if conn is None:
        return jsonify({"success": False, "error": "No hay conexión con la base de datos"}), 503

    archivo = request.files.get("file")
    if archivo is None or not archivo.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "error": "Selecciona un archivo .xlsx"}), 400

    try:
        consolidado = pd.read_excel(archivo, engine="openpyxl")

        def normalizar_columna(nombre):
            return unicodedata.normalize("NFKD", str(nombre)).encode("ascii", "ignore").decode().strip().lower()

        columnas = {normalizar_columna(nombre): nombre for nombre in consolidado.columns}

        def obtener_columna(*nombres):
            for nombre in nombres:
                if normalizar_columna(nombre) in columnas:
                    return columnas[normalizar_columna(nombre)]
            return None

        columna_sgi = obtener_columna("SGI")
        columna_inc = obtener_columna("Incremento Salarial", "inc_salarial")
        columna_promocion = obtener_columna("Promoción", "Promocion", "promocion")
        columna_nivelacion = obtener_columna("Nivelación", "Nivelacion", "nivelacion")
        if not all([columna_sgi, columna_inc, columna_promocion, columna_nivelacion]):
            return jsonify({"success": False, "error": "El consolidado no contiene las columnas requeridas"}), 400

        registros = []
        for _, fila in consolidado.iterrows():
            valor_sgi = fila[columna_sgi]
            sgi = str(int(valor_sgi)) if isinstance(valor_sgi, float) and valor_sgi.is_integer() else normalize_text(valor_sgi)
            if not sgi or sgi.lower() == "nan":
                continue
            inc = to_number(fila[columna_inc]) or 0
            promocion = to_number(fila[columna_promocion]) or 0
            nivelacion = to_number(fila[columna_nivelacion]) or 0
            registros.append((sgi, inc, promocion, nivelacion, inc + promocion + nivelacion))

        if not registros:
            return jsonify({"success": False, "error": "El consolidado no contiene SGI válidos"}), 400

        version_id = uuid.uuid4()
        columnas_budget = [
            "sgi", "nombre_completo", "compania", "puesto", "pais", "tipo_empleado",
            "salario_mensual", "inc_salarial", "promocion", "nivelacion", "suma_porcentual",
            "nuevo_salario", "fondo_ahorro", "aguinaldo", "prima_vacacional", "vales",
            "seguro_vida", "sgmm", "costo_total", "fecha_actualizacion", "version_id"
        ]
        columnas_sql = ", ".join(columnas_budget)

        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO budget_versiones (version_id, fecha_creacion, usuario, descripcion)
                VALUES (%s, NOW(), %s, %s)
            """, (version_id, "Sistema", "Actualización desde consolidado"))
            cur.execute(f"""
                INSERT INTO budget_actual_historico ({columnas_sql})
                SELECT {columnas_sql} FROM budget_actual
            """)

            actualizados = 0
            for sgi, inc, promocion, nivelacion, suma in registros:
                cur.execute("""
                    UPDATE budget_actual
                    SET inc_salarial = %s, promocion = %s, nivelacion = %s,
                        suma_porcentual = %s, nuevo_salario = salario_mensual * (1 + %s / 100),
                        fecha_actualizacion = NOW(), version_id = %s
                    WHERE sgi = %s
                """, (inc, promocion, nivelacion, suma, suma, version_id, sgi))
                actualizados += cur.rowcount

        conn.commit()
        return jsonify({"success": True, "ok": True, "version_id": str(version_id), "registros_actualizados": actualizados})
    except Exception as error:
        conn.rollback()
        print(f"Error actualizando sábana: {error}")
        return jsonify({"success": False, "error": str(error)}), 500


@app.route("/budget_versiones")
def budget_versiones():
    if conn is None:
        return jsonify([])
    with conn.cursor() as cur:
        cur.execute("""
            SELECT version_id, fecha_creacion, descripcion
            FROM budget_versiones ORDER BY fecha_creacion DESC
        """)
        versiones = cur.fetchall()
    return jsonify([
        {"version_id": str(version["version_id"]), "fecha_creacion": version["fecha_creacion"].isoformat(), "descripcion": version["descripcion"]}
        for version in versiones
    ])

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )

